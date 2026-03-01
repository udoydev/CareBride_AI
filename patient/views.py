import json
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Q
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render, reverse
from django.utils import timezone
from django.views.decorators.http import require_GET, require_POST

from carebridge.ai_services import GeminiAIService
from doctors.models import Appointment, DoctorSchedule
from prescriptions.models import FollowUp, Prescription, ReminderSchedule
from prescriptions.views import _build_prescription_pdf

from accounts.models import AppNotification, Doctor, Patient
from accounts.decorators import never_cache_auth

from .models import ChatMessage, ChatSession
from .services import generate_patient_reply, summarize_prescription


def _get_patient(request):
    if not request.user.is_authenticated:
        return None
    return getattr(request.user, "patient_profile", None)


def _resolve_language(request, patient):
    lang = request.GET.get("lang") or request.POST.get("lang")
    if not lang:
        try:
            if request.body:
                lang = json.loads(request.body).get("lang")
        except (json.JSONDecodeError, AttributeError, Exception):
            pass
    lang = lang or request.session.get("site_lang") or (patient.preferred_language if patient else "en")
    return "bn" if lang == "bn" else "en"


def _get_active_ai_model():
    try:
        from carebridge.ai_services import GeminiAIService
        providers = GeminiAIService._get_db_providers()
        for p in providers:
            if p.is_available:
                return p.model_name or p.get_provider_display()
    except Exception:
        pass
    return "Gemini"


def _get_or_create_session(patient, session_id=None):
    if not patient:
        return None
    if session_id:
        try:
            return ChatSession.objects.filter(pk=session_id, patient=patient).first()
        except (ValueError, TypeError):
            pass
    session = ChatSession.objects.create(patient=patient, title="New Chat")
    first_msg = ChatMessage.objects.filter(patient=patient, session=session).order_by("created_at").first()
    if first_msg:
        session.title = first_msg.content[:50] or "New Chat"
        session.save(update_fields=["title", "updated_at"])
    return session


DOSE_TIMES = [
    "09:00:00",
    "14:00:00",
    "18:00:00",
    "22:00:00",
]


def _ensure_dose_schedules(patient):
    from prescriptions.models import Prescription, PrescriptionItem, ReminderSchedule, get_active_dose_slots
    today = timezone.localdate()

    active_rxs = Prescription.objects.filter(patient=patient, status="active").order_by("-issued_at")
    if not active_rxs.exists():
        ReminderSchedule.objects.filter(
            prescription_item__prescription__patient=patient,
            scheduled_date=today,
            status="pending"
        ).delete()
        return

    latest_rx = active_rxs.first()
    older_rxs = active_rxs.exclude(pk=latest_rx.pk)
    if older_rxs.exists():
        older_rxs.update(status="completed")

    # Purge schedules for today belonging to non-latest prescriptions
    ReminderSchedule.objects.filter(
        prescription_item__prescription__patient=patient,
        scheduled_date=today,
    ).exclude(prescription_item__prescription=latest_rx).delete()

    items = latest_rx.items.all()
    for item in items:
        # Clean up legacy schedule entries without reminder_time
        ReminderSchedule.objects.filter(
            prescription_item=item,
            scheduled_date=today,
            reminder_time__isnull=True
        ).delete()

        issued_date = latest_rx.issued_at.date()
        days_since = (today - issued_date).days
        if 0 <= days_since < item.duration_days:
            active_slots = get_active_dose_slots(item.dosage, item.frequency)

            # Use patient custom dose times if available
            custom_times = patient.custom_dose_times or []
            if custom_times:
                # Map custom times to active slots, preserving labels from active_slots
                mapped_slots = []
                custom_idx = 0
                for slot in active_slots:
                    if custom_idx < len(custom_times):
                        new_slot = dict(slot)
                        new_slot["time"] = custom_times[custom_idx] + ":00" if len(custom_times[custom_idx]) == 5 else custom_times[custom_idx]
                        mapped_slots.append(new_slot)
                        custom_idx += 1
                active_slots = mapped_slots

            allowed_times = [slot["time"] for slot in active_slots]

            # Clean up any schedules for today that don't match active dosage slots
            ReminderSchedule.objects.filter(
                prescription_item=item,
                scheduled_date=today
            ).exclude(reminder_time__in=allowed_times).delete()

            for slot in active_slots:
                ReminderSchedule.objects.get_or_create(
                    prescription_item=item,
                    scheduled_date=today,
                    reminder_time=slot["time"],
                    defaults={"status": "pending"},
                )


@login_required
def custom_dose_times(request):
    patient = request.user.patient_profile

    if request.method == "POST":
        times = request.POST.getlist("dose_times")
        cleaned = []
        for t in times:
            t = t.strip()
            if not t:
                continue
            if len(t) == 5:
                t = t + ":00"
            cleaned.append(t)
        # Validate times
        valid_times = []
        for t in cleaned:
            try:
                from datetime import datetime
                datetime.strptime(t, "%H:%M:%S")
                valid_times.append(t[:5])
            except ValueError:
                pass
        patient.custom_dose_times = valid_times
        patient.save(update_fields=["custom_dose_times"])
        messages.success(request, "Custom dose times saved successfully.")
        return redirect("patient:doses_today")

    current_times = patient.custom_dose_times or []
    return render(request, "patient/custom_dose_times.html", {
        "current_times": current_times,
    })


@login_required
def dashboard(request):
    patient = request.user.patient_profile
    today = timezone.localdate()

    _ensure_dose_schedules(patient)

    now = timezone.localtime()
    current_hour = now.hour

    doses_today = ReminderSchedule.objects.filter(
        prescription_item__prescription__patient=patient,
        scheduled_date=today,
        status="pending",
    ).select_related("prescription_item__medicine")

    dose_reminder_message = None
    due_now = doses_today.filter(reminder_time__hour=current_hour).first()
    if due_now:
        dose_reminder_message = (
            f"Reminder sent to your mobile — "
            f"{due_now.prescription_item.medicine}, "
            f"{due_now.reminder_time.strftime('%I:%M %p')}"
        )

    next_follow_up = (
        FollowUp.objects.filter(prescription__patient=patient, status="upcoming")
        .order_by("scheduled_date")
        .first()
    )

    upcoming_appointments = Appointment.objects.filter(
        patient=patient,
        appointment_date__gte=timezone.localdate(),
        status__in=["pending", "confirmed"],
    ).select_related("doctor__user").order_by("appointment_date", "start_time")[:5]

    refund_notifications = AppNotification.objects.filter(
        user=request.user,
        notification_type="booking",
    ).filter(
        Q(title__icontains="refund") | Q(title__icontains="cancel") | Q(title__icontains="Cancellation")
    ).order_by("-created_at")[:5]

    return render(request, "patient/dashboard.html", {
        "doses_today": doses_today,
        "dose_reminder_message": dose_reminder_message,
        "next_follow_up": next_follow_up,
        "upcoming_appointments": upcoming_appointments,
        "refund_notifications": refund_notifications,
    })


@login_required
def prescription_detail(request, prescription_id):
    patient = request.user.patient_profile
    prescription = get_object_or_404(Prescription, pk=prescription_id, patient=patient)
    summary_language = request.GET.get("lang") or request.session.get("site_lang") or patient.preferred_language or "en"
    if summary_language not in {"bn", "en"}:
        summary_language = "en"
    
    if request.GET.get("lang"):
        request.session["site_lang"] = summary_language

    cache_key = f"prescription-summary:{prescription.pk}:{summary_language}"
    summary_payload = request.session.get(cache_key)
    if request.GET.get("refresh") or not summary_payload or not isinstance(summary_payload, dict) or "overview" not in summary_payload:
        summary_payload = summarize_prescription(prescription, summary_language)
        request.session[cache_key] = summary_payload

    # Deep AI analysis for detailed prescription view
    deep_cache_key = f"prescription-deep:{prescription.pk}:{summary_language}"
    deep_analysis = request.session.get(deep_cache_key)
    if request.GET.get("refresh") or not deep_analysis or not isinstance(deep_analysis, dict) or "overview" not in deep_analysis:
        from prescriptions.services import analyze_prescription_deep
        deep_analysis = analyze_prescription_deep(prescription, summary_language)
        request.session[deep_cache_key] = deep_analysis

    return render(request, "patient/prescription_detail.html", {
        "prescription": prescription,
        "summary_text": summary_payload.get("text", ""),
        "summary_overview": summary_payload.get("overview", ""),
        "summary_schedule": summary_payload.get("schedule", ""),
        "summary_precautions": summary_payload.get("precautions", ""),
        "summary_warnings": summary_payload.get("warnings", ""),
        "summary_source": summary_payload.get("source", "local"),
        "summary_language": summary_language,
        "deep_analysis": deep_analysis,
    })


@login_required
def doses_today(request):
    patient = request.user.patient_profile
    today = timezone.localdate()

    _ensure_dose_schedules(patient)

    schedules = ReminderSchedule.objects.filter(
        prescription_item__prescription__patient=patient,
        scheduled_date=today,
        status="pending",
    ).select_related("prescription_item__medicine")

    if request.method == "POST":
        schedule_id = request.POST.get("schedule_id")
        new_status = request.POST.get("status")
        schedule = get_object_or_404(
            ReminderSchedule, pk=schedule_id,
            prescription_item__prescription__patient=patient,
        )
        if new_status in ("taken", "skipped"):
            schedule.status = new_status
            schedule.save()
            messages.success(request, "✓ Dose status updated.")

        return redirect("patient:doses_today")

    return render(request, "patient/doses_today.html", {"schedules": schedules})


@login_required
def followups(request):
    patient = request.user.patient_profile
    today = timezone.localdate()

    all_followups = FollowUp.objects.filter(prescription__patient=patient).order_by("scheduled_date")

    # Auto-completion evaluator: if the doctor issued a prescription on or after follow-up date, mark completed
    for fu in all_followups:
        doc = fu.prescription.doctor
        has_rx = Prescription.objects.filter(
            doctor=doc,
            patient=patient,
            issued_at__date__gte=fu.scheduled_date
        ).exists()

        if has_rx and fu.status != "completed":
            fu.status = "completed"
            fu.save()
        elif fu.scheduled_date < today and fu.status == "upcoming":
            fu.status = "missed"
            fu.save()

    if request.method == "POST":
        followup_id = request.POST.get("followup_id")
        followup = get_object_or_404(FollowUp, pk=followup_id, prescription__patient=patient)
        followup.status = "completed"
        followup.save()
        messages.success(request, "✓ Follow-up marked as completed.")
        return redirect("patient:follow_ups")

    upcoming = all_followups.filter(status="upcoming")
    completed = all_followups.filter(status="completed")
    missed = all_followups.filter(status="missed")

    return render(request, "patient/follow_ups.html", {
        "upcoming": upcoming,
        "completed": completed,
        "missed": missed,
    })


@login_required
def notifications(request):
    from accounts.models import AppNotification
    notifications_qs = AppNotification.objects.filter(user=request.user)
    items = list(notifications_qs[:30])
    notifications_qs.filter(is_read=False).update(is_read=True)
    return render(request, "patient/notifications.html", {"items": items})


def _build_adherence_data(patient, weeks=4):
    from datetime import timedelta
    from collections import defaultdict

    today = timezone.localdate()
    current_week_start = today - timedelta(days=today.weekday())
    week_starts = sorted(
        current_week_start - timedelta(weeks=i)
        for i in range(weeks - 1, -1, -1)
    )
    week_labels = [w.strftime("%b %d") for w in week_starts]
    week_start_set = set(week_starts)

    schedules = ReminderSchedule.objects.filter(
        prescription_item__prescription__patient=patient,
        scheduled_date__gte=week_starts[0],
    ).select_related("prescription_item__medicine")

    medicine_stats = defaultdict(lambda: defaultdict(lambda: {"total": 0, "taken": 0}))
    for s in schedules:
        med_name = str(s.prescription_item.medicine)
        week_start = s.scheduled_date - timedelta(days=s.scheduled_date.weekday())
        if week_start in week_start_set:
            medicine_stats[med_name][week_start]["total"] += 1
            if s.status == "taken":
                medicine_stats[med_name][week_start]["taken"] += 1

    medicines = []
    for med_name in medicine_stats:
        data = []
        for w in week_starts:
            stats = medicine_stats[med_name].get(w, {"total": 0, "taken": 0})
            pct = round(stats["taken"] / stats["total"] * 100, 1) if stats["total"] > 0 else 0
            data.append(pct)
        medicines.append({"name": med_name, "data": data})

    return {
        "labels": week_labels,
        "medicines": medicines,
    }


@login_required
def health_record(request):
    from patient.models import PatientHealthReport

    patient = request.user.patient_profile
    language = request.GET.get("lang") or request.session.get("site_lang") or patient.preferred_language or "en"

    # 1. Handle Self-Report Upload
    if request.method == "POST" and "upload_report" in request.POST:
        title = (request.POST.get("title") or "").strip()
        report_type = request.POST.get("report_type", "lab_test")
        date_performed = request.POST.get("date_performed") or timezone.localdate()
        doc_clinic = (request.POST.get("doctor_or_clinic_name") or "").strip()
        result_desc = (request.POST.get("result_description") or "").strip()
        doc_file = request.FILES.get("document_file")

        if title and doc_file:
            PatientHealthReport.objects.create(
                patient=patient,
                title=title,
                report_type=report_type,
                date_performed=date_performed,
                doctor_or_clinic_name=doc_clinic,
                result_description=result_desc,
                document_file=doc_file,
            )
            messages.success(request, f"✓ Health report '{title}' uploaded successfully to your health journey.")
            return redirect("patient:health_record")
        else:
            messages.error(request, "Please provide a document title and file upload.")

    # 2. Fetch Prescriptions & Uploaded Health Reports
    prescriptions = (
        Prescription.objects.filter(patient=patient)
        .select_related("doctor__user")
        .prefetch_related("items__medicine", "follow_up")
        .order_by("-issued_at")
    )
    reports = PatientHealthReport.objects.filter(patient=patient).order_by("-date_performed")

    # 3. Build Unified Chronological Health Journey Timeline
    timeline_items = []
    for rx in prescriptions:
        timeline_items.append({
            "type": "prescription",
            "date": rx.issued_at.date(),
            "title": f"Prescription by Dr. {rx.doctor.user.get_full_name() or rx.doctor.user.username}",
            "subtitle": rx.doctor.specialty or "General Specialist",
            "details": f"Diagnosis: {rx.diagnosis or 'N/A'}\nComplaints: {rx.chief_complaints or 'N/A'}\nTests: {rx.tests_investigations or 'N/A'}\nAdvice: {rx.advice_rules or 'N/A'}",
            "obj": rx,
            "download_url": f"/patient/prescriptions/{rx.pk}/download/",
        })

    for r in reports:
        timeline_items.append({
            "type": "health_report",
            "date": r.date_performed,
            "title": r.title,
            "subtitle": f"{r.get_report_type_display()} &bull; {r.doctor_or_clinic_name or 'Self Uploaded'}",
            "details": r.result_description or "No result notes specified.",
            "obj": r,
            "file_url": r.document_file.url if r.document_file else "",
        })

    timeline_items.sort(key=lambda x: x["date"], reverse=True)

    # 4. Interactive AI Health Journey Assistant Q&A
    ai_answer = ""
    ai_query = (request.POST.get("ai_query") or "").strip() if (request.method == "POST" and "ai_query" in request.POST) else ""
    if ai_query:
        from patient.services import generate_patient_reply
        ai_answer = generate_patient_reply(
            question=ai_query,
            language=language,
            patient=patient,
            history=None,
            image_file=None,
            prescription_id=None,
        )

    # 5. Holistic AI Summary of Health History
    history_text_summary = f"Total Medical Records: {len(timeline_items)} ({len(prescriptions)} Doctor Prescriptions, {len(reports)} Self-Uploaded Lab Tests/External Reports)."

    adherence_chart = _build_adherence_data(patient)

    return render(request, "patient/health_record.html", {
        "prescriptions": prescriptions,
        "reports": reports,
        "timeline_items": timeline_items,
        "ai_answer": ai_answer,
        "ai_query": ai_query,
        "history_text_summary": history_text_summary,
        "ai_available": GeminiAIService.is_ai_available(),
        "adherence_chart": adherence_chart,
    })


@login_required
def doctor_list(request):
    from accounts.models import Doctor

    query = request.GET.get("q", "")
    category = request.GET.get("category", "")

    doctors = Doctor.objects.select_related("user").filter(is_verified=True)
    if query:
        doctors = doctors.filter(user__first_name__icontains=query) | doctors.filter(specialty__icontains=query)
    if category and category != "All":
        doctors = doctors.filter(specialty=category)

    categories = ["All"] + list(
        Doctor.objects.exclude(specialty="").values_list("specialty", flat=True).distinct()
    )

    return render(request, "patient/doctor_list.html", {
        "doctors": doctors,
        "categories": categories,
        "suggested": doctors[:2],
    })


@login_required
def doctor_detail(request, doctor_id):
    from accounts.models import Doctor
    from urllib.parse import quote_plus

    doctor = get_object_or_404(Doctor, pk=doctor_id)
    map_query = doctor.location_text.strip()
    map_url = f"https://www.google.com/maps/search/?api=1&query={quote_plus(map_query)}" if map_query else ""
    map_embed_url = f"https://www.google.com/maps?q={quote_plus(map_query)}&output=embed" if map_query else ""
    return render(request, "patient/doctor_detail.html", {
        "doctor": doctor,
        "map_url": map_url,
        "map_embed_url": map_embed_url,
    })


@login_required
def book_doctor(request, doctor_id):
    from doctors.models import DoctorSchedule, Appointment
    from accounts.models import Doctor, AppNotification

    doctor = get_object_or_404(Doctor, pk=doctor_id)
    patient = getattr(request.user, "patient_profile", None)
    if not patient:
        messages.error(request, "Only patients can book appointments.")
        return redirect("patient:doctor_list")

    schedules = DoctorSchedule.objects.filter(doctor=doctor, is_active=True).order_by("day_of_week", "start_time")

    if request.method == "POST":
        appointment_date = request.POST.get("appointment_date")
        start_time = request.POST.get("start_time")
        consultation_type = request.POST.get("consultation_type", "in_person")
        if consultation_type not in ["in_person", "video_online"]:
            consultation_type = "in_person"
        chief_complaint = request.POST.get("chief_complaint", "").strip()

        if not appointment_date or not start_time:
            messages.error(request, "Please select date and time.")
            return redirect("patient:book_doctor", doctor_id=doctor.pk)

        # Find matching schedule to calculate end_time
        from datetime import datetime, timedelta
        apt_date = datetime.strptime(appointment_date, "%Y-%m-%d").date()
        day_name = apt_date.strftime("%A").lower()
        schedule = DoctorSchedule.objects.filter(doctor=doctor, day_of_week=day_name, is_active=True).first()
        if not schedule:
            messages.error(request, "Doctor is not available on this day.")
            return redirect("patient:book_doctor", doctor_id=doctor.pk)

        start_dt = datetime.strptime(start_time, "%H:%M").time()
        end_dt = (datetime.combine(apt_date, start_dt) + timedelta(minutes=schedule.slot_duration_minutes)).time()

        # Double-booking check
        existing = Appointment.objects.filter(
            doctor=doctor,
            appointment_date=apt_date,
            status__in=["pending", "confirmed"],
        ).filter(start_time__lt=end_dt, end_time__gt=start_dt)
        if existing.exists():
            messages.error(request, "This time slot is already booked. Please choose another.")
            return redirect("patient:book_doctor", doctor_id=doctor.pk)

        appointment = Appointment.objects.create(
            patient=patient,
            doctor=doctor,
            appointment_date=apt_date,
            start_time=start_dt,
            end_time=end_dt,
            consultation_type=consultation_type,
            chief_complaint=chief_complaint,
            status="pending",
            fee_bdt=500.00,
            platform_fee_bdt=15.00,
            net_doctor_payout_bdt=485.00,
        )

        # Notification for doctor
        AppNotification.objects.create(
            user=doctor.user,
            title="🗓️ New Appointment Booking",
            message=f"Patient {request.user.get_full_name() or request.user.email} has booked an appointment on {apt_date} at {start_dt.strftime('%H:%M')}.",
            notification_type="booking",
            link_url=reverse("doctors:patient_detail", kwargs={"patient_id": patient.pk}),
        )

        # Notification for patient
        AppNotification.objects.create(
            user=request.user,
            title="✓ Appointment Booking Request Sent",
            message=f"Your booking request with Dr. {doctor.user.get_full_name() or doctor.user.username} on {apt_date} at {start_dt.strftime('%H:%M')} has been sent.",
            notification_type="booking",
            link_url=reverse("patient:notifications"),
        )

        messages.success(request, f"Appointment booked successfully for {apt_date} at {start_dt.strftime('%H:%M')}. Please complete payment.")
        return redirect("accounts:payment_process", appointment_id=appointment.pk)

    # Generate available slots for next 14 days
    from datetime import timedelta
    today = timezone.localdate()
    available_slots = []
    for i in range(14):
        date = today + timedelta(days=i)
        day_name = date.strftime("%A").lower()
        day_schedules = DoctorSchedule.objects.filter(doctor=doctor, day_of_week=day_name, is_active=True)
        for sch in day_schedules:
            slots = _generate_slots(sch.start_time, sch.end_time, sch.slot_duration_minutes, doctor, date)
            available_slots.extend(slots)

    return render(request, "patient/book_appointment.html", {
        "doctor": doctor,
        "schedules": schedules,
        "available_slots": available_slots,
    })


def _generate_slots(start_time, end_time, slot_duration, doctor, date):
    from datetime import datetime, timedelta
    slots = []
    current = datetime.combine(date, start_time)
    end = datetime.combine(date, end_time)
    while current + timedelta(minutes=slot_duration) <= end:
        slot_end = current + timedelta(minutes=slot_duration)
        is_booked = Appointment.objects.filter(
            doctor=doctor,
            appointment_date=date,
            status__in=["pending", "confirmed"],
            start_time=current.time(),
        ).exists()
        slots.append({
            "date": date.strftime("%Y-%m-%d"),
            "start": current.strftime("%H:%M"),
            "end": slot_end.strftime("%H:%M"),
            "available": not is_booked,
        })
        current = slot_end
    return slots


@login_required
def appointments(request):
    patient = getattr(request.user, "patient_profile", None)
    if not patient:
        messages.error(request, "Only patients can view appointments.")
        return redirect("home")

    status_filter = request.GET.get("status", "")
    apts_qs = Appointment.objects.filter(patient=patient).select_related("doctor__user").order_by("-appointment_date", "-start_time")
    if status_filter:
        apts_qs = apts_qs.filter(status=status_filter)

    now = timezone.now()
    apts = []
    for apt in apts_qs:
        apt_datetime = timezone.make_aware(
            timezone.datetime.combine(apt.appointment_date, apt.start_time)
        )
        hours_until = (apt_datetime - now).total_seconds() / 3600
        apt.can_cancel = apt.status in ("pending", "confirmed") and hours_until >= 24
        apt.can_edit = apt.status in ("pending", "confirmed") and hours_until >= 4 and apt.edit_count < 3
        apts.append(apt)

    return render(request, "patient/appointments.html", {
        "appointments": apts,
        "status_filter": status_filter,
    })


@login_required
def appointment_detail_patient(request, appointment_id):
    patient = getattr(request.user, "patient_profile", None)
    appointment = get_object_or_404(Appointment, pk=appointment_id, patient=patient)
    appointment_datetime = timezone.make_aware(
        timezone.datetime.combine(appointment.appointment_date, appointment.start_time)
    )
    hours_until = (appointment_datetime - timezone.now()).total_seconds() / 3600
    appointment.can_cancel = appointment.status in ("pending", "confirmed") and hours_until >= 24
    appointment.can_edit = appointment.status in ("pending", "confirmed") and hours_until >= 4 and appointment.edit_count < 3
    return render(request, "patient/appointment_detail.html", {"appointment": appointment})



from prescriptions.views import _build_prescription_pdf

@login_required
def download_prescription(request, prescription_id):
    patient = getattr(request.user, "patient_profile", None)
    prescription = get_object_or_404(Prescription, pk=prescription_id, patient=patient)
    buffer = _build_prescription_pdf(prescription)
    response = HttpResponse(buffer.read(), content_type="application/pdf")
    response["Content-Disposition"] = f"inline; filename=prescription_{prescription.pk}.pdf"
    return response


@login_required
def patient_analytics_view(request):
    patient = request.user.patient_profile
    today = timezone.localdate()

    appointments = Appointment.objects.filter(patient=patient)
    total_appointments = appointments.count()
    completed_appointments = appointments.filter(status="completed").count()
    pending_payments = appointments.filter(payment_status="pending").count()
    total_spent = appointments.filter(payment_status="paid").aggregate(total=Sum("fee_bdt"))["total"] or Decimal("0.00")

    upcoming_appointments = appointments.filter(
        appointment_date__gte=today, status__in=["pending", "confirmed"]
    ).count()
    prescriptions_count = Prescription.objects.filter(patient=patient).count()
    follow_ups_count = FollowUp.objects.filter(prescription__patient=patient, status="upcoming").count()

    recent_appointments = appointments.select_related("doctor__user").order_by("-appointment_date", "-start_time")[:10]
    recent_prescriptions = Prescription.objects.filter(patient=patient).select_related("doctor__user").order_by("-issued_at")[:10]

    return render(request, "patient/analytics.html", {
        "total_appointments": total_appointments,
        "completed_appointments": completed_appointments,
        "total_spent": total_spent,
        "pending_payments": pending_payments,
        "upcoming_appointments": upcoming_appointments,
        "prescriptions_count": prescriptions_count,
        "follow_ups_count": follow_ups_count,
        "recent_appointments": recent_appointments,
        "recent_prescriptions": recent_prescriptions,
    })


@login_required
def patient_payment_history(request):
    patient = request.user.patient_profile
    appointments = Appointment.objects.filter(patient=patient).select_related("doctor__user").order_by("-appointment_date", "-start_time")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payment History"

    headers = ["Date", "Doctor", "Specialty", "Fee (BDT)", "Platform Fee (BDT)", "Net Fee (BDT)", "Payment Status", "Transaction ID", "Consultation Type"]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for apt in appointments:
        ws.append([
            apt.appointment_date.strftime("%Y-%m-%d"),
            apt.doctor.user.get_full_name() or apt.doctor.user.username,
            apt.doctor.specialty or "General",
            float(apt.fee_bdt),
            float(apt.platform_fee_bdt),
            float(apt.net_doctor_payout_bdt),
            apt.get_payment_status_display(),
            apt.transaction_id or "N/A",
            apt.get_consultation_type_display(),
        ])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="payment_history.xlsx"'
    wb.save(response)
    return response


def chatbot(request):
    patient = _get_patient(request)
    prescription_id = request.GET.get("prescription_id") or request.POST.get("prescription_id")
    session_id = request.GET.get("session_id") or request.POST.get("session_id")
    
    if patient:
        session = _get_or_create_session(patient, session_id)
        qs = ChatMessage.objects.filter(patient=patient, session=session).order_by("created_at")[:50]
        formatted = [{"role": msg.role, "text": msg.content, "session": session.pk if session else None} for msg in qs]
        sessions = ChatSession.objects.filter(patient=patient).order_by("-updated_at")[:20]
    else:
        session = None
        request.session["guest_chat_messages"] = []
        request.session.modified = True
        formatted = []
        sessions = []

    if request.method == "POST" and patient:
        message = (request.POST.get("message") or "").strip()
        if message:
            language = _resolve_language(request, patient)
            ChatMessage.objects.create(patient=patient, session=session, role="user", content=message, language=language)
            history = list(ChatMessage.objects.filter(patient=patient, session=session).order_by("created_at"))
            reply = generate_patient_reply(
                question=message,
                language=language,
                patient=patient,
                history=history,
                prescription_id=prescription_id,
            )
            ChatMessage.objects.create(
                patient=patient,
                session=session,
                role="assistant",
                content=reply,
                language=language,
                ai_model_used=_get_active_ai_model(),
            )
            if session:
                session.title = session.messages.order_by("created_at").first().content[:50] if session.messages.exists() else "New Chat"
                session.save(update_fields=["title", "updated_at"])
            return redirect(f"{reverse('patient:chatbot')}?session_id={session.pk if session else ''}&prescription_id={prescription_id or ''}")

    guest_qa = []
    if not patient:
        from patient.services import GUEST_QA
        for key, qa in list(GUEST_QA.items())[:10]:
            guest_qa.append({
                "question_en": qa["en"].split("\n")[0].replace("**", "").strip(),
                "question_bn": qa["bn"].split("\n")[0].replace("**", "").strip(),
                "answer_en": qa["en"],
                "answer_bn": qa["bn"],
            })

    return render(request, "patient/chatbot.html", {
        "chat_messages": formatted,
        "sessions": sessions,
        "patient": patient,
        "prescription_id": prescription_id,
        "session_id": session_id,
        "active_session": session,
        "ai_available": GeminiAIService.is_ai_available() if patient else False,
        "active_model": _get_active_ai_model() if patient else None,
        "guest_qa": guest_qa,
    })


@require_GET
def chat_api_history(request):
    patient = _get_patient(request)
    if not patient:
        return JsonResponse({"messages": []})
    
    session_id = request.GET.get("session_id")
    
    if session_id:
        try:
            qs = ChatMessage.objects.filter(patient=patient, session_id=session_id).order_by("created_at")[:100]
        except (ValueError, TypeError):
            qs = ChatMessage.objects.filter(patient=patient).order_by("created_at")[:100]
    else:
        qs = ChatMessage.objects.filter(patient=patient).order_by("created_at")[:100]
    
    return JsonResponse({
        "messages": [
            {
                "id": msg.pk,
                "role": msg.role,
                "content": msg.content,
                "language": msg.language,
                "session": msg.session_id,
                "created_at": msg.created_at.isoformat(),
            }
            for msg in qs
        ],
    })


@require_POST
def chat_api_send(request):
    patient = _get_patient(request)
    if not patient:
        return JsonResponse({"error": "Login required"}, status=401)

    user_message = ""
    language = _resolve_language(request, patient)
    prescription_id = None
    session_id = None
    uploaded_file = request.FILES.get("file") or request.FILES.get("image") or request.FILES.get("prescription_image")

    if "application/json" in (request.content_type or "") and request.body:
        try:
            payload = json.loads(request.body.decode("utf-8"))
            user_message = (payload.get("message") or "").strip()
            language = payload.get("lang") or language
            prescription_id = payload.get("prescription_id")
            session_id = payload.get("session_id")
        except (json.JSONDecodeError, AttributeError):
            pass
    else:
        user_message = (request.POST.get("message") or "").strip()
        language = request.POST.get("lang") or language
        prescription_id = request.POST.get("prescription_id")
        session_id = request.POST.get("session_id")

    if not user_message and not uploaded_file:
        return JsonResponse({"error": "Message or uploaded document is required."}, status=400)

    user_content = user_message
    if uploaded_file:
        user_content += f" 📎 [Attached File: {uploaded_file.name}]"

    if patient:
        session = _get_or_create_session(patient, session_id)
        ChatMessage.objects.create(
            patient=patient,
            session=session,
            role="user",
            content=user_content,
            language=language,
        )

        history = list(ChatMessage.objects.filter(patient=patient, session=session).order_by("created_at"))
        reply = generate_patient_reply(
            question=user_message,
            language=language,
            patient=patient,
            history=history,
            image_file=uploaded_file,
            prescription_id=prescription_id,
        )

        assistant_msg = ChatMessage.objects.create(
            patient=patient,
            session=session,
            role="assistant",
            content=reply,
            language=language,
            ai_model_used=_get_active_ai_model(),
        )

        if session:
            session.title = session.messages.order_by("created_at").first().content[:50] if session.messages.exists() else "New Chat"
            session.save(update_fields=["title", "updated_at"])

        return JsonResponse({
            "reply": reply,
            "message_id": assistant_msg.pk,
            "session_id": session.pk if session else None,
            "language": language,
        })
    else:
        request.session["guest_chat_messages"] = []
        request.session.modified = True
        session_messages = []
        session_messages.append({"role": "user", "content": user_content, "language": language})
        
        reply = generate_patient_reply(
            question=user_message,
            language=language,
            patient=None,
            history=session_messages,
            image_file=uploaded_file,
            prescription_id=prescription_id,
        )
        session_messages.append({"role": "assistant", "content": reply, "language": language})
        
        request.session["guest_chat_messages"] = session_messages[-5:]
        request.session.modified = True

        return JsonResponse({
            "reply": reply,
            "message_id": len(session_messages),
            "session_id": None,
            "language": language,
        })



@require_POST
def chat_api_clear(request):
    patient = _get_patient(request)
    if not patient:
        return JsonResponse({"error": "Login required"}, status=401)
    
    session_id = request.POST.get("session_id")
    if session_id:
        try:
            ChatMessage.objects.filter(patient=patient, session_id=session_id).delete()
            ChatSession.objects.filter(pk=session_id, patient=patient).delete()
        except (ValueError, TypeError):
            pass
    else:
        ChatMessage.objects.filter(patient=patient).delete()
        ChatSession.objects.filter(patient=patient).delete()
    
    return JsonResponse({"ok": True})


@require_POST
def chat_api_new_session(request):
    patient = _get_patient(request)
    if not patient:
        return JsonResponse({"error": "Login required"}, status=401)
    session = ChatSession.objects.create(patient=patient, title="New Chat")
    return JsonResponse({
        "session_id": session.pk,
        "title": session.title,
        "created_at": session.created_at.isoformat(),
    })


@require_GET
def chat_api_sessions(request):
    patient = _get_patient(request)
    if not patient:
        return JsonResponse({"sessions": []})
    sessions = ChatSession.objects.filter(patient=patient).order_by("-updated_at")[:50]
    data = []
    for s in sessions:
        last_msg = s.messages.order_by("-created_at").first()
        data.append({
            "id": s.pk,
            "title": s.title,
            "updated_at": s.updated_at.isoformat(),
            "last_message": last_msg.content[:60] if last_msg else "",
            "message_count": s.messages.count(),
        })
    return JsonResponse({"sessions": data})


def chat_ui(request):
    """Gemini-like chat interface with file upload and doctor suggestions."""
    patient = _get_patient(request)
    language = _resolve_language(request, patient)
    prescription_id = request.GET.get("prescription_id") or request.POST.get("prescription_id")
    session_id = request.GET.get("session_id") or request.POST.get("session_id")

    if request.method == "POST":
        message = (request.POST.get("message") or "").strip()
        uploaded_file = request.FILES.get("file") or request.FILES.get("image") or request.FILES.get("prescription_image")

        if not message and not uploaded_file:
            return JsonResponse({"error": "Message or file required."}, status=400)

        user_content = message
        file_url = None
        if uploaded_file:
            user_content = message or f"Analyze this document: {uploaded_file.name}"
            file_url = uploaded_file.name

        if not patient:
            return JsonResponse({"error": "Login required for AI chat."}, status=401)

        session = _get_or_create_session(patient, session_id)
        ChatMessage.objects.create(
            patient=patient,
            session=session,
            role="user",
            content=user_content,
            language=language,
        )
        history = list(ChatMessage.objects.filter(patient=patient, session=session).order_by("created_at"))
        reply = generate_patient_reply(
            question=message,
            language=language,
            patient=patient,
            history=history,
            image_file=uploaded_file,
            prescription_id=prescription_id,
        )
        assistant_msg = ChatMessage.objects.create(
            patient=patient,
            session=session,
            role="assistant",
            content=reply,
            language=language,
            ai_model_used=_get_active_ai_model(),
        )
        if session:
            session.title = session.messages.order_by("created_at").first().content[:50] if session.messages.exists() else "New Chat"
            session.save(update_fields=["title", "updated_at"])
        return JsonResponse({
            "reply": reply,
            "message_id": assistant_msg.pk,
            "session_id": session.pk if session else None,
            "language": language,
            "file_url": file_url,
        })

    # GET request — render chat UI
    active_session = None
    if patient:
        if session_id:
            try:
                active_session = ChatSession.objects.filter(pk=session_id, patient=patient).first()
            except (ValueError, TypeError):
                pass
        if not active_session:
            active_session = ChatSession.objects.filter(patient=patient).order_by("-updated_at").first()
        if active_session:
            messages = ChatMessage.objects.filter(patient=patient, session=active_session).order_by("created_at")[:100]
            formatted = [{"role": msg.role, "text": msg.content, "id": msg.pk, "language": msg.language, "created_at": msg.created_at} for msg in messages]
        else:
            formatted = []
        sessions = ChatSession.objects.filter(patient=patient).order_by("-updated_at")[:50]
    else:
        session_messages = request.session.get("guest_chat_messages", [])
        formatted = [{"role": msg["role"], "text": msg["content"], "language": msg.get("language", "en")} for msg in session_messages]
        sessions = []
        active_session = None

    suggested_doctors = []
    if patient:
        from accounts.models import Doctor
        suggested_doctors = list(Doctor.objects.filter(is_verified=True).select_related("user")[:8])

    return render(request, "patient/chat_ui.html", {
        "chat_messages": formatted,
        "patient": patient,
        "language": language,
        "sessions": sessions,
        "active_session": active_session,
        "suggested_doctors": suggested_doctors,
        "prescription_id": prescription_id,
        "ai_available": GeminiAIService.is_ai_available(),
        "active_model": _get_active_ai_model(),
    })


@login_required
def request_cancellation(request, appointment_id):
    patient = request.user.patient_profile
    appointment = get_object_or_404(Appointment, pk=appointment_id, patient=patient)

    if appointment.status not in ("confirmed", "pending"):
        messages.error(request, "This appointment cannot be cancelled.")
        return redirect("patient:appointments")

    if request.method == "POST":
        reason = request.POST.get("reason", "").strip()
        appointment_datetime = timezone.make_aware(
            timezone.datetime.combine(appointment.appointment_date, appointment.start_time)
        )
        hours_until = (appointment_datetime - timezone.now()).total_seconds() / 3600

        if hours_until >= 24:
            appointment.status = "cancelled"
            appointment.cancellation_reason = reason
            appointment.refund_status = "partial"
            appointment.refund_amount = (appointment.fee_bdt * Decimal("0.35")).quantize(Decimal("0.01"))
            appointment.net_doctor_payout_bdt = appointment.fee_bdt - appointment.platform_fee_bdt - appointment.refund_amount
            appointment.save(update_fields=["status", "cancellation_reason", "refund_status", "refund_amount", "net_doctor_payout_bdt"])

            AppNotification.objects.create(
                user=appointment.doctor.user,
                title="Appointment Cancelled by Patient",
                message=f"Patient {patient.user.get_full_name()} cancelled appointment on {appointment.appointment_date}. Refund: {appointment.refund_amount} BDT (35%) processed.",
                notification_type="booking",
                link_url=reverse("doctors:appointment_list"),
            )
            AppNotification.objects.create(
                user=request.user,
                title="Cancellation Confirmed",
                message=f"Your appointment on {appointment.appointment_date} has been cancelled. Refund: {appointment.refund_amount} BDT (35%) has been processed.",
                notification_type="booking",
                link_url=reverse("patient:appointments"),
            )
            messages.success(request, f"Appointment cancelled. {appointment.refund_amount} BDT refunded (35%).")
        else:
            appointment.status = "cancellation_pending"
            appointment.cancellation_reason = reason
            appointment.cancellation_requested_at = timezone.now()
            appointment.save(update_fields=["status", "cancellation_reason", "cancellation_requested_at"])

            AppNotification.objects.create(
                user=appointment.doctor.user,
                title="Cancellation Request from Patient",
                message=f"Patient {patient.user.get_full_name()} requested cancellation for {appointment.appointment_date}. Please review. Refund policy: 35% if approved.",
                notification_type="booking",
                link_url=reverse("doctors:appointment_list"),
            )
            AppNotification.objects.create(
                user=request.user,
                title="Cancellation Request Sent",
                message=f"Your cancellation request for {appointment.appointment_date} has been sent to the doctor for approval.",
                notification_type="booking",
                link_url=reverse("patient:appointments"),
            )
            messages.info(request, "Cancellation request sent to doctor for approval. You will be notified once decided.")

        return redirect("patient:appointments")

    return render(request, "patient/request_cancellation.html", {"appointment": appointment})


@login_required
def edit_appointment(request, appointment_id):
    patient = request.user.patient_profile
    appointment = get_object_or_404(Appointment, pk=appointment_id, patient=patient)

    if appointment.status not in ("pending", "confirmed"):
        messages.error(request, "Only pending or confirmed appointments can be edited.")
        return redirect("patient:appointments")

    if appointment.edit_count >= 3:
        messages.error(request, "You have reached the maximum of 3 edits for this booking.")
        return redirect("patient:appointments")

    appointment_datetime = timezone.make_aware(
        timezone.datetime.combine(appointment.appointment_date, appointment.start_time)
    )
    hours_until = (appointment_datetime - timezone.now()).total_seconds() / 3600
    if hours_until < 4:
        messages.error(request, "Appointments can only be edited at least 4 hours before the scheduled time.")
        return redirect("patient:appointments")

    if request.method == "POST":
        new_date = request.POST.get("appointment_date", "").strip()
        new_start_time = request.POST.get("start_time", "").strip()
        new_consultation_type = request.POST.get("consultation_type", appointment.consultation_type)

        if not new_date or not new_start_time:
            messages.error(request, "Please select both date and time.")
            return redirect("patient:edit_appointment", appointment_id=appointment.pk)

        if new_consultation_type not in ["in_person", "video_online"]:
            new_consultation_type = "in_person"

        from datetime import datetime, timedelta
        apt_date = datetime.strptime(new_date, "%Y-%m-%d").date()
        day_name = apt_date.strftime("%A").lower()
        schedule = DoctorSchedule.objects.filter(doctor=appointment.doctor, day_of_week=day_name, is_active=True).first()
        if not schedule:
            messages.error(request, "Doctor is not available on this day.")
            return redirect("patient:edit_appointment", appointment_id=appointment.pk)

        start_dt = datetime.strptime(new_start_time, "%H:%M").time()
        end_dt = (datetime.combine(apt_date, start_dt) + timedelta(minutes=schedule.slot_duration_minutes)).time()

        existing = Appointment.objects.filter(
            doctor=appointment.doctor,
            appointment_date=apt_date,
            status__in=["pending", "confirmed"],
        ).filter(start_time__lt=end_dt, end_time__gt=start_dt).exclude(pk=appointment.pk)
        if existing.exists():
            messages.error(request, "This time slot is already booked. Please choose another.")
            return redirect("patient:edit_appointment", appointment_id=appointment.pk)

        old_date = appointment.appointment_date
        old_time = appointment.start_time.strftime("%H:%M")
        appointment.appointment_date = apt_date
        appointment.start_time = start_dt
        appointment.end_time = end_dt
        appointment.consultation_type = new_consultation_type
        appointment.edit_count += 1
        appointment.save(update_fields=["appointment_date", "start_time", "end_time", "consultation_type", "edit_count"])

        AppNotification.objects.create(
            user=appointment.doctor.user,
            title="Appointment Rescheduled by Patient",
            message=f"Patient {patient.user.get_full_name()} changed appointment from {old_date} {old_time} to {apt_date} {start_dt.strftime('%H:%M')}. Edit count: {appointment.edit_count}/3.",
            notification_type="booking",
            link_url=reverse("doctors:appointment_list"),
        )
        AppNotification.objects.create(
            user=request.user,
            title="Appointment Updated",
            message=f"Your appointment with Dr. {appointment.doctor.user.get_full_name()} has been updated to {apt_date} {start_dt.strftime('%H:%M')}.",
            notification_type="booking",
            link_url=reverse("patient:appointments"),
        )

        messages.success(request, f"Appointment updated successfully. Edit count: {appointment.edit_count}/3.")
        return redirect("patient:appointments")

    schedules = DoctorSchedule.objects.filter(doctor=appointment.doctor, is_active=True).order_by("day_of_week", "start_time")
    available_slots = []
    from datetime import timedelta
    today = timezone.localdate()
    for i in range(14):
        date = today + timedelta(days=i)
        day_name = date.strftime("%A").lower()
        day_schedules = DoctorSchedule.objects.filter(doctor=appointment.doctor, day_of_week=day_name, is_active=True)
        for sch in day_schedules:
            slots = _generate_slots(sch.start_time, sch.end_time, sch.slot_duration_minutes, appointment.doctor, date)
            available_slots.extend(slots)

    return render(request, "patient/edit_appointment.html", {
        "appointment": appointment,
        "doctor": appointment.doctor,
        "schedules": schedules,
        "available_slots": available_slots,
        "edit_count": appointment.edit_count,
    })
