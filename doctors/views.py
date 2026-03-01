from decimal import Decimal
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render, reverse
from django.utils import timezone

from accounts.decorators import never_cache_auth
from accounts.models import AppNotification, Doctor, Patient
from carebridge.ai_services import GeminiAIService
from doctors.models import Appointment, DoctorSchedule
from prescriptions.models import FollowUp, Medicine, Prescription, PrescriptionItem, ReminderSchedule


@never_cache_auth
@login_required
def dashboard(request):
    doctor = getattr(request.user, "doctor_profile", None)
    if not doctor:
        return render(request, "doctors/dashboard.html", {"rows": []})

    today = timezone.localdate()

    # Automatic Follow-up Status Evaluator
    followups = FollowUp.objects.filter(prescription__doctor=doctor)
    for fu in followups:
        has_new_prescription = Prescription.objects.filter(
            doctor=doctor,
            patient=fu.prescription.patient,
            issued_at__date__gte=fu.scheduled_date
        ).exists()

        if has_new_prescription and fu.status != "completed":
            fu.status = "completed"
            fu.save()
        elif fu.scheduled_date < today and fu.status == "upcoming":
            fu.status = "missed"
            fu.save()

    patient_rows = []
    prescriptions = (
        Prescription.objects.filter(doctor=doctor)
        .select_related("patient__user")
        .prefetch_related("items__medicine", "follow_up")
        .order_by("-issued_at")
    )

    seen_patient_ids = set()
    for prescription in prescriptions:
        patient = prescription.patient
        if patient.pk in seen_patient_ids:
            continue
        seen_patient_ids.add(patient.pk)
        follow_up = getattr(prescription, "follow_up", None)
        adherence = None
        if follow_up:
            if follow_up.status == "completed":
                adherence = 100
            elif follow_up.status == "missed":
                adherence = 40
            else:
                adherence = 85

        patient_rows.append({
            "patient": {
                "id": patient.pk,
                "name": patient.user.get_full_name() or patient.user.email,
                "avatar": patient.avatar.url if patient.avatar else None,
                "district": patient.district or "Dhaka",
            },
            "adherence": adherence,
            "follow_up": follow_up,
        })

    # Stats
    total_patients = len(seen_patient_ids)
    today_appointments = Appointment.objects.filter(doctor=doctor, appointment_date=today).count()
    pending_appointments = Appointment.objects.filter(doctor=doctor, status="pending").count()
    upcoming_followups = FollowUp.objects.filter(prescription__doctor=doctor, status="upcoming").count()

    stats = {
        "total_patients": total_patients,
        "today_appointments": today_appointments,
        "pending_appointments": pending_appointments,
        "upcoming_followups": upcoming_followups,
    }

    refund_notifications = AppNotification.objects.filter(
        user=request.user,
        notification_type="booking",
    ).filter(
        Q(title__icontains="refund") | Q(title__icontains="cancel") | Q(title__icontains="Cancellation")
    ).order_by("-created_at")[:5]

    doctor_name = doctor.user.get_full_name() or doctor.user.email
    has_any_prescription = Prescription.objects.filter(doctor=doctor).exists()
    return render(request, "doctors/dashboard.html", {
        "rows": patient_rows,
        "today": today,
        "stats": stats,
        "doctor_name": doctor_name,
        "doctor": doctor,
        "has_any_prescription": has_any_prescription,
        "refund_notifications": refund_notifications,
    })


@never_cache_auth
@login_required
def patient_list(request):
    doctor = getattr(request.user, "doctor_profile", None)
    query = request.GET.get("q", "").strip()

    patients_qs = Patient.objects.select_related("user").filter(is_verified=True)
    if query:
        patients_qs = patients_qs.filter(
            Q(user__first_name__icontains=query)
            | Q(user__last_name__icontains=query)
            | Q(user__email__icontains=query)
            | Q(phone_number__icontains=query)
        )

    patients = [
        {
            "id": p.pk,
            "name": p.user.get_full_name() or p.user.email,
            "district": p.district or "Dhaka",
            "phone": p.phone_number or "N/A",
            "avatar": p.avatar.url if p.avatar else None,
        }
        for p in patients_qs[:50]
    ]

    return render(request, "doctors/patient_list.html", {"patients": patients, "query": query})


@never_cache_auth
@login_required
def update_followup_status(request, followup_id):
    doctor = getattr(request.user, "doctor_profile", None)
    follow_up = get_object_or_404(FollowUp, pk=followup_id, prescription__doctor=doctor)
    new_status = request.GET.get("status") or request.POST.get("status")
    if new_status in {"completed", "missed", "upcoming"}:
        follow_up.status = new_status
        follow_up.save()
        messages.success(request, f"✓ Follow-up status updated to '{new_status.capitalize()}'.")
    return redirect("doctors:dashboard")


@never_cache_auth
@login_required
def patient_detail(request, patient_id):
    doctor = getattr(request.user, "doctor_profile", None)
    patient = get_object_or_404(Patient, pk=patient_id)
    prescription_qs = (
        Prescription.objects.filter(patient=patient)
        .select_related("patient__user", "doctor__user")
        .prefetch_related("items__medicine", "follow_up")
        .order_by("-issued_at")
    )
    prescriptions = []
    for prescription in prescription_qs:
        prescriptions.append({
            "id": prescription.pk,
            "date": prescription.issued_at.strftime("%Y-%m-%d"),
            "status": prescription.get_status_display(),
            "doctor": prescription.doctor.user.get_full_name() or prescription.doctor.user.username,
            "items": [f"{item.medicine} — {item.dosage} ({item.frequency}x/day)" for item in prescription.items.all()],
            "follow_up": getattr(prescription, "follow_up", None),
        })

    patient_name = patient.user.get_full_name() or patient.user.email
    ai_summary = GeminiAIService.generate_clinical_summary(
        patient_name=patient_name,
        history_text="Patient has active prescription and follow-up records on CareBridge.",
        metrics_summary="Data pulled from CareBridge prescription history.",
    )

    from patient.models import PatientHealthReport
    reports = PatientHealthReport.objects.filter(patient=patient).order_by("-date_performed")

    return render(request, "doctors/patient_detail.html", {
        "patient": {"id": patient.pk, "name": patient_name, "avatar": patient.avatar.url if patient.avatar else None},
        "prescriptions": prescriptions,
        "reports": reports,
        "adherence": 100 if prescriptions else None,
        "ai_summary": ai_summary,
    })


@never_cache_auth
@login_required
def create_prescription(request, patient_id):
    doctor = getattr(request.user, "doctor_profile", None)
    if not doctor:
        messages.error(request, "Only verified doctor profiles can write prescriptions.")
        return redirect("doctors:dashboard")

    patient = get_object_or_404(Patient, pk=patient_id)

    if request.method == "POST":
        chief_complaints = request.POST.get("chief_complaints", "").strip()
        diagnosis = request.POST.get("diagnosis", "").strip()
        doctor_notes = request.POST.get("doctor_notes", "").strip()
        follow_up_date_str = request.POST.get("follow_up_date", "").strip()

        # Dynamic lists
        med_names = request.POST.getlist("med_name[]")
        med_dosages = request.POST.getlist("med_dosage[]")
        med_frequencies = request.POST.getlist("med_frequency[]")
        med_timings = request.POST.getlist("med_timing[]")
        med_durations = request.POST.getlist("med_duration[]")
        med_notes_list = request.POST.getlist("med_notes[]")

        test_names = [t.strip() for t in request.POST.getlist("test_name[]") if t.strip()]
        advice_rules = [a.strip() for a in request.POST.getlist("advice_rule[]") if a.strip()]

        fu_date = None
        if follow_up_date_str:
            try:
                fu_date = timezone.datetime.strptime(follow_up_date_str, "%Y-%m-%d").date()
            except ValueError:
                fu_date = None

        # Supersede / complete previous active prescriptions for this patient
        Prescription.objects.filter(patient=patient, status="active").update(status="completed")

        # Create Prescription Instance
        prescription = Prescription.objects.create(
            doctor=doctor,
            patient=patient,
            chief_complaints=chief_complaints,
            diagnosis=diagnosis,
            tests_investigations="\n".join(test_names),
            advice_rules="\n".join(advice_rules),
            doctor_notes=doctor_notes,
            next_followup_date=fu_date,
            status="active",
        )

        # Loop and save medicines
        for i, name in enumerate(med_names):
            name_clean = name.strip()
            if not name_clean:
                continue
            dosage = med_dosages[i].strip() if i < len(med_dosages) and med_dosages[i].strip() else "1 tablet"
            try:
                freq = int(med_frequencies[i]) if i < len(med_frequencies) else 2
            except (ValueError, TypeError):
                freq = 2
            timing = med_timings[i].strip() if i < len(med_timings) and med_timings[i].strip() else "after_meal"
            try:
                duration = int(med_durations[i]) if i < len(med_durations) else 7
            except (ValueError, TypeError):
                duration = 7
            notes = med_notes_list[i].strip() if i < len(med_notes_list) else ""

            med_obj, _ = Medicine.objects.get_or_create(
                brand_name=name_clean,
                defaults={"generic_name": name_clean, "form": "tablet"}
            )

            item_obj = PrescriptionItem.objects.create(
                prescription=prescription,
                medicine=med_obj,
                dosage=dosage,
                frequency=freq,
                timing_relation_to_meal=timing,
                duration_days=duration,
                special_instructions=notes,
            )

            # Auto-generate daily dose schedules for the full duration based on exact dosage notation
            from prescriptions.models import get_active_dose_slots
            today_date = timezone.localdate()
            active_slots = get_active_dose_slots(dosage, freq)
            for day_offset in range(max(1, duration)):
                sch_date = today_date + timezone.timedelta(days=day_offset)
                for slot in active_slots:
                    ReminderSchedule.objects.get_or_create(
                        prescription_item=item_obj,
                        scheduled_date=sch_date,
                        reminder_time=slot["time"],
                        defaults={"status": "pending"}
                    )

        # Auto-complete any existing upcoming follow-ups for this patient with this doctor
        FollowUp.objects.filter(
            prescription__doctor=doctor,
            prescription__patient=patient,
            status="upcoming"
        ).update(status="completed")

        # Auto-complete any pending/confirmed appointments for this doctor+patient on the appointment date
        Appointment.objects.filter(
            doctor=doctor,
            patient=patient,
            status__in=["pending", "confirmed"],
        ).update(status="completed")

        # Save new FollowUp object if date selected
        if fu_date:
            FollowUp.objects.create(
                prescription=prescription,
                scheduled_date=fu_date,
                status="upcoming",
            )

        messages.success(request, f"✓ Medical Prescription #{prescription.pk} issued successfully for {patient.user.get_full_name() or patient.user.email}!")
        return redirect("doctors:download_prescription", prescription_id=prescription.pk)

    return render(request, "doctors/create_prescription.html", {
        "patient": patient,
        "patient_name": patient.user.get_full_name() or patient.user.email,
        "doctor": doctor,
    })


@never_cache_auth
@login_required
def notifications(request):
    from accounts.models import AppNotification
    notifications_qs = AppNotification.objects.filter(user=request.user)
    items = list(notifications_qs[:30])
    notifications_qs.filter(is_read=False).update(is_read=True)
    return render(request, "doctors/notifications.html", {"items": items})


@never_cache_auth
@login_required
def history(request):
    doctor = getattr(request.user, "doctor_profile", None)
    activity = []
    if doctor:
        for prescription in Prescription.objects.filter(doctor=doctor).select_related("patient__user").order_by("-issued_at")[:15]:
            activity.append({
                "date": prescription.issued_at.strftime("%Y-%m-%d"),
                "action": f"Issued prescription for {prescription.patient.user.get_full_name() or prescription.patient.user.email}",
            })
    return render(request, "doctors/history.html", {"activity": activity})


@never_cache_auth
@login_required
def profile_edit(request):
    doctor = getattr(request.user, "doctor_profile", None)
    categories = ["General Physician", "Cardiologist", "Dermatologist", "Pediatrician", "Orthopedic", "Gastroenterologist", "Neurologist", "ENT Specialist"]

    if request.method == "POST":
        full_name = request.POST.get("full_name", "").strip()
        specialty = request.POST.get("specialty", "General Physician").strip()
        clinic_name = request.POST.get("clinic_name", "").strip()
        location_text = request.POST.get("location_text", "").strip()
        bio = request.POST.get("bio", "").strip()
        experience_years = request.POST.get("experience_years", "").strip()
        consultation_fee = request.POST.get("consultation_fee", "").strip()

        errors = []
        if not experience_years:
            errors.append("Experience years is required.")
        if not consultation_fee:
            errors.append("Consultation fee is required.")

        if full_name:
            names = full_name.split(" ", 1)
            request.user.first_name = names[0]
            request.user.last_name = names[1] if len(names) > 1 else ""
            request.user.save()

        if doctor:
            doctor.specialty = specialty
            doctor.clinic_name = clinic_name
            doctor.location_text = location_text
            doctor.bio = bio
            try:
                doctor.experience_years = int(experience_years) if experience_years else 0
            except (ValueError, TypeError):
                errors.append("Experience years must be a valid number.")
            try:
                doctor.consultation_fee = Decimal(consultation_fee) if consultation_fee else Decimal("0")
            except (ValueError, TypeError):
                errors.append("Consultation fee must be a valid amount.")

            if request.FILES.get("avatar"):
                old_avatar = doctor.avatar
                doctor.avatar = request.FILES.get("avatar")
                doctor.avatar_updated_at = timezone.now()
                if old_avatar and old_avatar.name != doctor.avatar.name:
                    old_avatar.delete(save=False)
            doctor.save()

        if errors:
            for error in errors:
                messages.error(request, error)
            return redirect("doctors:profile_edit")

        messages.success(request, "✓ Doctor profile updated successfully.")
        return redirect("doctors:profile_edit")

    doctor_payload = {
        "name": request.user.get_full_name() or request.user.email,
        "specialty": doctor.specialty if doctor else "General Physician",
        "bio": doctor.bio if doctor else "",
        "clinic_name": doctor.clinic_name if doctor else "",
        "location_text": doctor.location_text if doctor else "",
        "experience_years": doctor.experience_years if doctor else 0,
        "consultation_fee": doctor.consultation_fee if doctor else 0,
        "avatar": doctor.avatar.url if (doctor and doctor.avatar) else None,
        "avatar_updated_at": doctor.avatar_updated_at if doctor else None,
    }
    return render(request, "doctors/profile_edit.html", {"doctor": doctor_payload, "categories": categories})


@never_cache_auth
@login_required
def download_prescription(request, prescription_id):
    from prescriptions.views import _build_prescription_pdf
    prescription = get_object_or_404(Prescription, pk=prescription_id)
    buffer = _build_prescription_pdf(prescription)
    response = HttpResponse(buffer.read(), content_type="application/pdf")
    response["Content-Disposition"] = f"inline; filename=prescription_{prescription.pk}.pdf"
    return response


@never_cache_auth
@login_required
def prescription_detail(request, prescription_id):
    prescription = get_object_or_404(Prescription, pk=prescription_id)
    summary_language = request.GET.get("lang") or request.session.get("site_lang") or "en"
    if summary_language not in {"bn", "en"}:
        summary_language = "en"

    from patient.services import summarize_prescription
    summary_payload = summarize_prescription(prescription, summary_language)

    return render(request, "patient/prescription_detail.html", {
        "prescription": prescription,
        "summary_text": summary_payload.get("text", ""),
        "summary_overview": summary_payload.get("overview", ""),
        "summary_schedule": summary_payload.get("schedule", ""),
        "summary_precautions": summary_payload.get("precautions", ""),
        "summary_warnings": summary_payload.get("warnings", ""),
        "summary_source": summary_payload.get("source", "local"),
        "summary_language": summary_language,
    })


@never_cache_auth
@login_required
def schedule_management(request):
    doctor = getattr(request.user, "doctor_profile", None)
    if not doctor:
        messages.error(request, "Only doctors can manage schedules.")
        return redirect("home")

    if request.method == "POST":
        day = request.POST.get("day_of_week")
        start = request.POST.get("start_time")
        end = request.POST.get("end_time")
        slot = request.POST.get("slot_duration_minutes", 20)
        if day and start and end:
            DoctorSchedule.objects.create(
                doctor=doctor,
                day_of_week=day,
                start_time=start,
                end_time=end,
                slot_duration_minutes=int(slot),
                is_active=True,
            )
            messages.success(request, "Schedule added successfully.")
        return redirect("doctors:schedule_management")

    schedules = DoctorSchedule.objects.filter(doctor=doctor).order_by("day_of_week", "start_time")
    days = DoctorSchedule.DAY_CHOICES
    return render(request, "doctors/schedule_management.html", {"schedules": schedules, "days": days})


@never_cache_auth
@login_required
def delete_schedule(request, schedule_id):
    doctor = getattr(request.user, "doctor_profile", None)
    schedule = get_object_or_404(DoctorSchedule, pk=schedule_id, doctor=doctor)
    schedule.delete()
    messages.success(request, "Schedule removed.")
    return redirect("doctors:schedule_management")


@never_cache_auth
@login_required
def appointment_list(request):
    doctor = getattr(request.user, "doctor_profile", None)
    if not doctor:
        messages.error(request, "Only doctors can view appointments.")
        return redirect("home")

    status_filter = request.GET.get("status", "")
    today = timezone.localdate()

    base_qs = Appointment.objects.filter(doctor=doctor).select_related("patient__user")
    if status_filter == "today":
        appointments = base_qs.filter(appointment_date=today).order_by("-appointment_date", "-start_time")
    elif status_filter == "pending":
        appointments = base_qs.filter(status="pending").order_by("-appointment_date", "-start_time")
    elif status_filter == "missed":
        appointments = base_qs.filter(status="missed").order_by("-appointment_date", "-start_time")
    elif status_filter:
        appointments = base_qs.filter(status=status_filter).order_by("-appointment_date", "-start_time")
    else:
        appointments = base_qs.order_by("-appointment_date", "-start_time")

    # Stats
    all_appointments = base_qs
    stats = {
        "total": all_appointments.count(),
        "pending": all_appointments.filter(status="pending").count(),
        "confirmed": all_appointments.filter(status="confirmed").count(),
        "completed": all_appointments.filter(status="completed").count(),
        "missed": all_appointments.filter(status="missed").count(),
        "cancelled": all_appointments.filter(status="cancelled").count(),
        "cancellation_pending": all_appointments.filter(status="cancellation_pending").count(),
        "today": all_appointments.filter(appointment_date=today).count(),
    }

    return render(request, "doctors/appointment_list.html", {
        "appointments": appointments,
        "status_filter": status_filter,
        "stats": stats,
        "today": today,
    })


@never_cache_auth
@login_required
def appointment_detail(request, appointment_id):
    doctor = getattr(request.user, "doctor_profile", None)
    appointment = get_object_or_404(Appointment, pk=appointment_id, doctor=doctor)

    if request.method == "POST":
        appointment.status = request.POST.get("status", appointment.status)
        appointment.notes = request.POST.get("notes", appointment.notes)
        appointment.save()
        messages.success(request, "Appointment updated.")
        return redirect("doctors:appointment_detail", appointment_id=appointment.pk)

    return render(request, "doctors/appointment_detail.html", {"appointment": appointment})


@never_cache_auth
@login_required
def send_emergency_notification(request):
    doctor = getattr(request.user, "doctor_profile", None)
    if not doctor:
        messages.error(request, "Only doctors can send notifications.")
        return redirect("home")

    if request.method == "POST":
        title = request.POST.get("title", "").strip()
        message = request.POST.get("message", "").strip()
        appointment_date = request.POST.get("appointment_date", "").strip()
        send_all = request.POST.get("send_all") == "on"

        if not title or not message:
            messages.error(request, "Title and message are required.")
            return redirect("doctors:send_emergency_notification")

        affected_appointments = Appointment.objects.filter(doctor=doctor, status="confirmed")
        if appointment_date:
            affected_appointments = affected_appointments.filter(appointment_date=appointment_date)
        elif not send_all:
            today = timezone.localdate()
            affected_appointments = affected_appointments.filter(appointment_date__gte=today)

        patient_users = set(affected_appointments.values_list("patient__user", flat=True))
        for user_id in patient_users:
            AppNotification.objects.create(
                user_id=user_id,
                title=title,
                message=message,
                notification_type="booking",
                link_url=reverse("patient:notifications"),
            )

        messages.success(request, f"Emergency notification sent to {len(patient_users)} patients.")
        return redirect("doctors:appointment_list")

    today = timezone.localdate()
    upcoming_count = Appointment.objects.filter(doctor=doctor, status="confirmed", appointment_date__gte=today).count()
    return render(request, "doctors/send_emergency.html", {
        "upcoming_count": upcoming_count,
        "today": today,
    })


@never_cache_auth
@login_required
def doctor_financial_report(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    doctor = getattr(request.user, "doctor_profile", None)
    if not doctor:
        messages.error(request, "Only doctors can view financial reports.")
        return redirect("home")

    appointments = Appointment.objects.filter(doctor=doctor).select_related("patient__user").order_by("-appointment_date", "-start_time")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financial Report"

    headers = ["Date", "Patient", "Fee (BDT)", "Payment Status", "Platform Fee (BDT)", "Net Payout (BDT)", "Consultation Type"]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for apt in appointments:
        ws.append([
            apt.appointment_date.strftime("%Y-%m-%d"),
            apt.patient.user.get_full_name() or apt.patient.user.username,
            float(apt.fee_bdt),
            apt.get_payment_status_display(),
            float(apt.platform_fee_bdt),
            float(apt.net_doctor_payout_bdt),
            apt.get_consultation_type_display(),
        ])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="financial_report.xlsx"'
    wb.save(response)
    return response


@never_cache_auth
@login_required
def approve_cancellation(request, appointment_id):
    doctor = getattr(request.user, "doctor_profile", None)
    if not doctor:
        messages.error(request, "Access restricted to doctors.")
        return redirect("home")

    appointment = get_object_or_404(Appointment, pk=appointment_id, doctor=doctor)

    if appointment.status != "cancellation_pending":
        messages.error(request, "This appointment does not have a pending cancellation.")
        return redirect("doctors:appointment_detail", appointment_id=appointment.pk)

    if request.method == "POST":
        action = request.POST.get("action")
        reason = request.POST.get("reason", "").strip()

        if action == "approve":
            appointment.status = "cancelled"
            appointment.cancellation_approved = True
            appointment.refund_status = "partial"
            appointment.refund_amount = (appointment.fee_bdt * Decimal("0.35")).quantize(Decimal("0.01"))
            appointment.net_doctor_payout_bdt = appointment.fee_bdt - appointment.platform_fee_bdt - appointment.refund_amount
            appointment.save(update_fields=["status", "cancellation_approved", "refund_status", "refund_amount", "net_doctor_payout_bdt"])

            AppNotification.objects.create(
                user=appointment.patient.user,
                title="Cancellation Approved",
                message=f"Your cancellation for {appointment.appointment_date} was approved. Refund: {appointment.refund_amount} BDT (35%) has been processed.",
                notification_type="booking",
                link_url=reverse("patient:appointments"),
            )
            AppNotification.objects.create(
                user=request.user,
                title="Cancellation Approved — Refund Issued",
                message=f"You approved cancellation for {appointment.patient.user.get_full_name()} on {appointment.appointment_date}. Patient refunded {appointment.refund_amount} BDT (35%). Your payout adjusted.",
                notification_type="booking",
                link_url=reverse("doctors:appointment_list"),
            )
            messages.success(request, f"Cancellation approved. Patient refunded {appointment.refund_amount} BDT (35%).")
        elif action == "reject":
            appointment.status = "cancelled"
            appointment.cancellation_approved = False
            appointment.refund_status = "full"
            appointment.refund_amount = appointment.fee_bdt
            appointment.net_doctor_payout_bdt = 0
            appointment.save(update_fields=["status", "cancellation_approved", "refund_status", "refund_amount", "net_doctor_payout_bdt"])

            AppNotification.objects.create(
                user=appointment.patient.user,
                title="Cancellation Rejected — Full Refund Issued",
                message=f"Your cancellation request for {appointment.appointment_date} was rejected. As a result, a full refund of {appointment.refund_amount} BDT has been issued to your account.",
                notification_type="booking",
                link_url=reverse("patient:appointments"),
            )
            AppNotification.objects.create(
                user=request.user,
                title="Cancellation Rejected — Full Refund to Patient",
                message=f"You rejected cancellation for {appointment.patient.user.get_full_name()} on {appointment.appointment_date}. Full refund of {appointment.refund_amount} BDT issued to patient.",
                notification_type="booking",
                link_url=reverse("doctors:appointment_list"),
            )
            messages.info(request, f"Cancellation rejected. Full refund of {appointment.refund_amount} BDT issued to patient.")

        return redirect("doctors:appointment_detail", appointment_id=appointment.pk)

    return render(request, "doctors/approve_cancellation.html", {"appointment": appointment})


@never_cache_auth
@login_required
def mark_attendance(request, appointment_id):
    doctor = getattr(request.user, "doctor_profile", None)
    if not doctor:
        messages.error(request, "Access restricted to doctors.")
        return redirect("home")

    appointment = get_object_or_404(Appointment, pk=appointment_id, doctor=doctor)

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "visited":
            appointment.status = "completed"
            appointment.save(update_fields=["status"])
            messages.success(request, f"Marked {appointment.patient.user.get_full_name()} as visited.")
        elif action == "missed":
            appointment.status = "missed"
            appointment.save(update_fields=["status"])

            AppNotification.objects.create(
                user=appointment.patient.user,
                title="Appointment Missed",
                message=f"You missed your appointment on {appointment.appointment_date.strftime('%d %b %Y')} with Dr. {request.user.get_full_name()}. Please reschedule if needed.",
                notification_type="booking",
                link_url=reverse("patient:doctor_list"),
            )
            messages.warning(request, f"Marked {appointment.patient.user.get_full_name()} as missed. Notification sent.")

    return redirect("doctors:appointment_list")


@never_cache_auth
@login_required
def auto_detect_missed(request):
    doctor = getattr(request.user, "doctor_profile", None)
    if not doctor:
        messages.error(request, "Access restricted to doctors.")
        return redirect("home")

    today = timezone.localdate()
    cutoff = today

    missed_appointments = Appointment.objects.filter(
        doctor=doctor,
        appointment_date__lt=cutoff,
        status__in=["confirmed", "pending"],
    ).select_related("patient__user")

    count = 0
    for apt in missed_appointments:
        has_prescription = Prescription.objects.filter(
            doctor=doctor,
            patient=apt.patient,
            issued_at__date__gte=apt.appointment_date,
        ).exists()

        if not has_prescription:
            apt.status = "missed"
            apt.save(update_fields=["status"])

            AppNotification.objects.create(
                user=apt.patient.user,
                title="Appointment Marked as Missed",
                message=f"Your appointment on {apt.appointment_date.strftime('%d %b %Y')} with Dr. {request.user.get_full_name()} was marked as missed (no prescription generated). Please book again if needed.",
                notification_type="booking",
                link_url=reverse("patient:doctor_list"),
            )
            count += 1

    messages.success(request, f"Auto-detected and marked {count} appointment(s) as missed.")
    return redirect("doctors:appointment_list")


@never_cache_auth
@login_required
def appointment_report(request):
    doctor = getattr(request.user, "doctor_profile", None)
    if not doctor:
        messages.error(request, "Access restricted to doctors.")
        return redirect("home")

    today = timezone.localdate()
    appointments = Appointment.objects.filter(doctor=doctor).select_related("patient__user")

    total = appointments.count()
    visited = appointments.filter(status="completed").count()
    missed = appointments.filter(status="missed").count()
    cancelled = appointments.filter(status="cancelled").count()
    pending = appointments.filter(status="pending").count()
    confirmed = appointments.filter(status="confirmed").count()

    visit_rate = (visited / total * 100) if total > 0 else 0
    missed_rate = (missed / total * 100) if total > 0 else 0
    cancelled_rate = (cancelled / total * 100) if total > 0 else 0

    report_data = {
        "total": total,
        "visited": visited,
        "missed": missed,
        "cancelled": cancelled,
        "pending": pending,
        "confirmed": confirmed,
        "visit_rate": round(visit_rate, 1),
        "missed_rate": round(missed_rate, 1),
        "cancelled_rate": round(cancelled_rate, 1),
        "appointments": appointments.order_by("-appointment_date", "-start_time")[:100],
    }

    return render(request, "doctors/appointment_report.html", report_data)


@never_cache_auth
@login_required
def appointment_report_export(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    doctor = getattr(request.user, "doctor_profile", None)
    if not doctor:
        messages.error(request, "Access restricted to doctors.")
        return redirect("home")

    appointments = Appointment.objects.filter(doctor=doctor).select_related("patient__user").order_by("-appointment_date", "-start_time")

    total = appointments.count()
    visited = appointments.filter(status="completed").count()
    missed = appointments.filter(status="missed").count()
    cancelled = appointments.filter(status="cancelled").count()
    pending = appointments.filter(status="pending").count()
    confirmed = appointments.filter(status="confirmed").count()

    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["CareBridge AI - Doctor Appointment Report"])
    ws_summary.append([f"Doctor: Dr. {request.user.get_full_name()}"])
    ws_summary.append([f"Generated: {timezone.localdate().strftime('%d %b %Y')}"])
    ws_summary.append([])
    ws_summary.append(["Metric", "Count", "Percentage"])
    ws_summary.append(["Total Appointments", total, "100%"])
    ws_summary.append(["Visited / Completed", visited, f"{round(visited/total*100, 1) if total > 0 else 0}%"])
    ws_summary.append(["Missed / No-Show", missed, f"{round(missed/total*100, 1) if total > 0 else 0}%"])
    ws_summary.append(["Cancelled", cancelled, f"{round(cancelled/total*100, 1) if total > 0 else 0}%"])
    ws_summary.append(["Pending", pending, f"{round(pending/total*100, 1) if total > 0 else 0}%"])
    ws_summary.append(["Confirmed", confirmed, f"{round(confirmed/total*100, 1) if total > 0 else 0}%"])

    for cell in ws_summary[5]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    ws_detail = wb.create_sheet("Appointment Details")
    headers = ["Date", "Patient Name", "Phone", "Status", "Consultation Type", "Fee (BDT)", "Payment Status", "Chief Complaint"]
    ws_detail.append(headers)
    for cell in ws_detail[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for apt in appointments:
        ws_detail.append([
            apt.appointment_date.strftime("%Y-%m-%d"),
            apt.patient.user.get_full_name() or apt.patient.user.username,
            apt.patient.phone_number or "N/A",
            apt.get_status_display(),
            apt.get_consultation_type_display(),
            float(apt.fee_bdt),
            apt.get_payment_status_display(),
            (apt.chief_complaint or "")[:100],
        ])

    for ws in [ws_summary, ws_detail]:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="appointment_report.xlsx"'
    wb.save(response)
    return response


